import streamlit as st
import pandas as pd
import re
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
#  PAGE CONFIG
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="RME Data Cleaner",
    page_icon="🏥",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────
#  CUSTOM CSS
# ─────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700;800&family=JetBrains+Mono:wght@400;500&display=swap');

html, body, [class*="css"] {
    font-family: 'Plus Jakarta Sans', sans-serif;
}

/* ── Background ── */
.stApp {
    background: linear-gradient(135deg, #0f1623 0%, #1a2744 50%, #0f1623 100%);
    min-height: 100vh;
}

/* ── Sidebar ── */
section[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #111827 0%, #1e2d4a 100%);
    border-right: 1px solid rgba(99,179,237,0.15);
}
section[data-testid="stSidebar"] * { color: #e2e8f0 !important; }
section[data-testid="stSidebar"] .stSelectbox label,
section[data-testid="stSidebar"] .stCheckbox label { color: #94a3b8 !important; }

/* ── Header ── */
.main-header {
    background: linear-gradient(135deg, rgba(30,58,138,0.6) 0%, rgba(14,165,233,0.15) 100%);
    border: 1px solid rgba(99,179,237,0.3);
    border-radius: 16px;
    padding: 2rem 2.5rem;
    margin-bottom: 1.5rem;
    position: relative;
    overflow: hidden;
}
.main-header::before {
    content: '';
    position: absolute;
    top: -50%;
    right: -10%;
    width: 300px;
    height: 300px;
    background: radial-gradient(circle, rgba(14,165,233,0.12) 0%, transparent 70%);
    border-radius: 50%;
}
.main-header h1 {
    color: #f0f9ff;
    font-size: 2.2rem;
    font-weight: 800;
    margin: 0 0 0.25rem 0;
    letter-spacing: -0.03em;
}
.main-header p {
    color: #7dd3fc;
    font-size: 0.95rem;
    margin: 0;
    font-weight: 500;
}

/* ── Metric cards ── */
.metric-row {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 1rem;
    margin-bottom: 1.5rem;
}
.metric-card {
    background: rgba(30,41,59,0.8);
    border: 1px solid rgba(99,179,237,0.2);
    border-radius: 12px;
    padding: 1.25rem 1.5rem;
    text-align: center;
    backdrop-filter: blur(8px);
}
.metric-card .label {
    color: #64748b;
    font-size: 0.75rem;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    margin-bottom: 0.5rem;
}
.metric-card .value {
    color: #e2e8f0;
    font-size: 1.9rem;
    font-weight: 800;
    line-height: 1;
    font-family: 'JetBrains Mono', monospace;
}
.metric-card .value.good { color: #34d399; }
.metric-card .value.warn { color: #fbbf24; }
.metric-card .value.bad  { color: #f87171; }

/* ── Section card ── */
.section-card {
    background: rgba(15,22,35,0.7);
    border: 1px solid rgba(99,179,237,0.15);
    border-radius: 14px;
    padding: 1.5rem;
    margin-bottom: 1rem;
    backdrop-filter: blur(6px);
}
.section-title {
    color: #7dd3fc;
    font-size: 0.8rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.1em;
    margin-bottom: 1rem;
    display: flex;
    align-items: center;
    gap: 0.5rem;
}

/* ── Issue badges ── */
.badge {
    display: inline-block;
    padding: 0.2rem 0.65rem;
    border-radius: 999px;
    font-size: 0.72rem;
    font-weight: 600;
    font-family: 'JetBrains Mono', monospace;
}
.badge-red   { background: rgba(248,113,113,0.15); color: #f87171; border: 1px solid rgba(248,113,113,0.3); }
.badge-yellow{ background: rgba(251,191,36,0.12);  color: #fbbf24; border: 1px solid rgba(251,191,36,0.3); }
.badge-green { background: rgba(52,211,153,0.12);  color: #34d399; border: 1px solid rgba(52,211,153,0.3); }
.badge-blue  { background: rgba(99,179,237,0.12);  color: #63b3ed; border: 1px solid rgba(99,179,237,0.3); }

/* ── Log item ── */
.log-item {
    display: flex;
    align-items: flex-start;
    gap: 0.6rem;
    padding: 0.6rem 0;
    border-bottom: 1px solid rgba(99,179,237,0.07);
    font-size: 0.85rem;
    color: #94a3b8;
}
.log-icon { font-size: 1rem; flex-shrink: 0; margin-top: 1px; }

/* ── Upload zone ── */
.upload-hint {
    background: rgba(14,165,233,0.06);
    border: 2px dashed rgba(14,165,233,0.25);
    border-radius: 12px;
    padding: 1.5rem;
    text-align: center;
    color: #475569;
    font-size: 0.9rem;
    margin-bottom: 1rem;
}

/* ── Merge card ── */
.merge-card {
    background: linear-gradient(135deg, rgba(20,30,50,0.9) 0%, rgba(10,20,40,0.9) 100%);
    border: 1px solid rgba(168,85,247,0.3);
    border-radius: 14px;
    padding: 1.5rem;
    margin-bottom: 1rem;
}
.merge-title {
    color: #c084fc;
    font-size: 0.8rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.1em;
    margin-bottom: 1rem;
}
.file-tag {
    display: inline-flex;
    align-items: center;
    gap: 0.4rem;
    background: rgba(168,85,247,0.1);
    border: 1px solid rgba(168,85,247,0.25);
    border-radius: 8px;
    padding: 0.35rem 0.75rem;
    font-size: 0.78rem;
    color: #c084fc;
    font-family: 'JetBrains Mono', monospace;
    margin: 0.2rem;
}
.file-tag.sorted {
    background: rgba(52,211,153,0.1);
    border-color: rgba(52,211,153,0.25);
    color: #34d399;
}

/* ── Dataframe ── */
[data-testid="stDataFrame"] { border-radius: 10px; overflow: hidden; }

/* ── Buttons ── */
.stButton > button {
    background: linear-gradient(135deg, #1d4ed8, #0ea5e9) !important;
    color: white !important;
    border: none !important;
    border-radius: 10px !important;
    font-weight: 700 !important;
    font-family: 'Plus Jakarta Sans', sans-serif !important;
    font-size: 0.9rem !important;
    padding: 0.6rem 1.5rem !important;
    transition: all 0.2s !important;
    box-shadow: 0 4px 15px rgba(14,165,233,0.25) !important;
}
.stButton > button:hover {
    transform: translateY(-1px) !important;
    box-shadow: 0 6px 20px rgba(14,165,233,0.4) !important;
}

/* ── Download button ── */
.stDownloadButton > button {
    background: linear-gradient(135deg, #065f46, #10b981) !important;
    color: white !important;
    border: none !important;
    border-radius: 10px !important;
    font-weight: 700 !important;
    font-family: 'Plus Jakarta Sans', sans-serif !important;
    width: 100% !important;
    padding: 0.75rem !important;
    font-size: 1rem !important;
    box-shadow: 0 4px 15px rgba(16,185,129,0.3) !important;
}

/* ── Checkboxes & selects ── */
.stCheckbox > label { color: #cbd5e1 !important; font-size: 0.9rem !important; }
.stSelectbox > label { color: #94a3b8 !important; font-size: 0.8rem !important; font-weight: 600 !important; }
.stMultiSelect > label { color: #94a3b8 !important; font-size: 0.8rem !important; font-weight: 600 !important; }

/* ── Divider ── */
hr { border-color: rgba(99,179,237,0.1) !important; }

/* ── Expander ── */
.streamlit-expanderHeader {
    background: rgba(30,41,59,0.5) !important;
    border-radius: 8px !important;
    color: #7dd3fc !important;
}

/* ── Tab styling ── */
.stTabs [data-baseweb="tab-list"] {
    background: rgba(15,22,35,0.7) !important;
    border-radius: 12px !important;
    padding: 0.3rem !important;
    border: 1px solid rgba(99,179,237,0.15) !important;
    gap: 0.3rem !important;
}
.stTabs [data-baseweb="tab"] {
    border-radius: 8px !important;
    color: #64748b !important;
    font-weight: 600 !important;
    font-size: 0.9rem !important;
}
.stTabs [aria-selected="true"] {
    background: linear-gradient(135deg, #1d4ed8, #0ea5e9) !important;
    color: white !important;
}
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────
#  CONSTANTS
# ─────────────────────────────────────────────
BULAN_MAP = {
    "januari": 1, "februari": 2, "maret": 3, "april": 4,
    "mei": 5, "juni": 6, "juli": 7, "agustus": 8,
    "september": 9, "oktober": 10, "november": 11, "desember": 12,
    "jan": 1, "feb": 2, "mar": 3, "apr": 4,
    "may": 5, "jun": 6, "jul": 7, "aug": 8,
    "sep": 9, "oct": 10, "nov": 11, "dec": 12,
    "january": 1, "february": 2, "march": 3,
    "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10
}

BULAN_LABEL = {
    1: "Januari", 2: "Februari", 3: "Maret", 4: "April",
    5: "Mei", 6: "Juni", 7: "Juli", 8: "Agustus",
    9: "September", 10: "Oktober", 11: "November", 12: "Desember"
}


# ─────────────────────────────────────────────
#  HELPERS — DATA QUALITY
# ─────────────────────────────────────────────

def detect_issues(df: pd.DataFrame) -> dict:
    issues = {}
    trailing_comma_cols = []
    for col in df.select_dtypes(include="object").columns:
        if df[col].dropna().astype(str).str.contains(r'\s*,\s*$', regex=True).any():
            trailing_comma_cols.append(col)
    if trailing_comma_cols:
        issues["trailing_comma"] = trailing_comma_cols

    dup_count = df.duplicated().sum()
    if dup_count:
        issues["duplicates"] = int(dup_count)

    nulls = df.isnull().sum()
    null_cols = nulls[nulls > 0].to_dict()
    if null_cols:
        issues["nulls"] = null_cols

    placeholder_cols = []
    for col in df.select_dtypes(include="object").columns:
        if df[col].dropna().astype(str).str.fullmatch(r'[-=]+').any():
            placeholder_cols.append(col)
    if placeholder_cols:
        issues["placeholders"] = placeholder_cols

    date_cols = []
    for col in df.columns:
        if "tgl" in col.lower() or "tanggal" in col.lower() or "date" in col.lower() or "lahir" in col.lower():
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                date_cols.append(col)
    if date_cols:
        issues["date_cols"] = date_cols

    return issues


def clean_dataframe(df: pd.DataFrame, options: dict) -> tuple[pd.DataFrame, list]:
    df = df.copy()
    log = []

    if options.get("remove_duplicates"):
        before = len(df)
        df.drop_duplicates(inplace=True)
        removed = before - len(df)
        log.append(("🗑️", f"Hapus {removed} baris duplikat") if removed else ("✅", "Tidak ada baris duplikat"))

    if options.get("strip_trailing_comma"):
        affected = []
        for col in df.select_dtypes(include="object").columns:
            mask = df[col].dropna().astype(str).str.contains(r'\s*,\s*$', regex=True)
            if mask.any():
                df[col] = df[col].astype(str).str.replace(r'\s*,\s*$', '', regex=True).str.strip()
                affected.append(col)
        if affected:
            log.append(("✂️", f"Hapus trailing koma pada: {', '.join(affected)}"))

    if options.get("strip_spaces"):
        for col in options.get("strip_spaces_cols", []):
            if col in df.columns:
                df[col] = df[col].astype(str).str.replace(r'\s+', '', regex=True)
        log.append(("🔤", "Hapus spasi berlebih di dalam NIK / No Penjamin"))

    if options.get("format_dates"):
        fmt = options.get("date_format", "%d/%m/%Y")
        for col in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                df[col] = pd.to_datetime(df[col]).dt.strftime(fmt)
                log.append(("📅", f"Format tanggal kolom '{col}' → {fmt}"))

    if options.get("standardize_keterangan") and "Keterangan" in df.columns:
        placeholder_val = options.get("placeholder_replacement", "Tidak Ada Keterangan")
        df["Keterangan"] = df["Keterangan"].astype(str).str.strip()
        df["Keterangan"] = df["Keterangan"].replace(["-", "=", "nan", ""], placeholder_val)
        df["Keterangan"] = df["Keterangan"].apply(
            lambda x: placeholder_val if re.fullmatch(r'[-=\s]+', str(x)) else x
        )
        if options.get("uppercase_keterangan"):
            df["Keterangan"] = df["Keterangan"].str.upper()
        log.append(("📝", f"Standarisasi Keterangan: '-', '=' → '{placeholder_val}'"))

    if options.get("uppercase_nama") and "Nama" in df.columns:
        df["Nama"] = df["Nama"].str.upper().str.strip()
        log.append(("🔠", "Nama diubah ke UPPERCASE"))

    if options.get("titlecase_desa") and "Desa" in df.columns:
        df["Desa"] = df["Desa"].str.strip().str.title()
        log.append(("🏘️", "Desa diubah ke Title Case"))

    fill_map = options.get("fill_nulls", {})
    for col, fill_val in fill_map.items():
        if col in df.columns and fill_val:
            n = int(df[col].isnull().sum())
            if n:
                df[col] = df[col].fillna(fill_val)
                log.append(("📋", f"Isi {n} nilai kosong di '{col}' dengan '{fill_val}'"))

    rename_map = options.get("rename_cols", {})
    if rename_map:
        df.rename(columns=rename_map, inplace=True)
        for old, new in rename_map.items():
            log.append(("✏️", f"Rename kolom '{old}' → '{new}'"))

    drop_cols = [c for c in options.get("drop_cols", []) if c in df.columns]
    if drop_cols:
        df.drop(columns=drop_cols, inplace=True)
        log.append(("❌", f"Hapus kolom: {', '.join(drop_cols)}"))

    sort_col = options.get("sort_by")
    if sort_col and sort_col in df.columns:
        df.sort_values(sort_col, inplace=True)
        df.reset_index(drop=True, inplace=True)
        log.append(("🔃", f"Urutkan berdasarkan '{sort_col}'"))

    return df, log


def to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Data") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    h_fill  = PatternFill('solid', start_color='1F4E79')
    h_font  = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
    h_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    alt_fill = PatternFill('solid', start_color='EBF3FB')
    d_font   = Font(name='Calibri', size=9)
    d_align  = Alignment(vertical='center')
    thin = Side(style='thin', color='B0C4DE')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font = h_font; c.fill = h_fill
        c.alignment = h_align; c.border = bdr

    df_str = df.fillna('-')
    for ri, row in enumerate(df_str.itertuples(index=False), 2):
        fill = alt_fill if ri % 2 == 0 else PatternFill('solid', start_color='FFFFFF')
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = d_font; c.fill = fill
            c.alignment = d_align; c.border = bdr

    for ci, col in enumerate(df.columns, 1):
        max_len = max(
            len(str(col)),
            df_str.iloc[:, ci-1].astype(str).str.len().max() if len(df_str) else 0
        )
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 3, 45)

    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 32

    ws2 = wb.create_sheet('Info')
    summary = [
        ['File diekspor oleh RME Data Cleaner'],
        ['Total baris:', len(df)],
        ['Total kolom:', len(df.columns)],
    ]
    for ri, row in enumerate(summary, 1):
        for ci, val in enumerate(row, 1):
            ws2.cell(row=ri, column=ci, value=val)
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────
#  HELPERS — MERGE
# ─────────────────────────────────────────────

def extract_period_from_filename(filename: str) -> dict:
    """
    Ekstrak bulan dan tahun dari nama file.
    Mengembalikan dict dengan keys: month (int|None), year (int|None), label (str)
    """
    name = filename.lower()
    name_clean = re.sub(r'[_\-\s\.]+', ' ', name)

    month = None
    year = None

    # Cari tahun (4 digit: 2000–2099)
    year_match = re.search(r'\b(20\d{2})\b', name_clean)
    if year_match:
        year = int(year_match.group(1))

    # Cari bulan dari nama
    for bname, bnum in BULAN_MAP.items():
        pattern = r'\b' + re.escape(bname) + r'\b'
        if re.search(pattern, name_clean):
            month = bnum
            break

    # Cari bulan dari angka (mis. 01, 02, ... 12) jika belum ketemu
    if month is None:
        # Pola: angka 01-12 yang berdiri sendiri atau dipisah underscore/dash
        m = re.search(r'(?<!\d)(0?[1-9]|1[0-2])(?!\d)', name_clean)
        if m:
            month = int(m.group(1))

    label_parts = []
    if month:
        label_parts.append(BULAN_LABEL.get(month, str(month)))
    if year:
        label_parts.append(str(year))
    label = " ".join(label_parts) if label_parts else filename

    return {"month": month, "year": year, "label": label}


def sort_key_period(info: dict) -> tuple:
    return (info["year"] or 9999, info["month"] or 99)


def merge_excel_files(
    uploaded_files,
    add_source_col: bool = True,
    add_month_col: bool = False,
    source_col_name: str = "Sumber",
    month_col_name: str = "Bulan",
    sheet_per_month: bool = False,
    add_summary_sheet: bool = True,
) -> tuple[bytes, list]:
    """
    Gabungkan beberapa file Excel/CSV menjadi satu file.
    Diurutkan otomatis berdasarkan bulan/tahun dari nama file.
    Mengembalikan (excel_bytes, merge_log).
    """
    log = []
    parsed = []

    for f in uploaded_files:
        info = extract_period_from_filename(f.name)
        parsed.append({"file": f, "info": info})

    # Urutkan berdasarkan tahun lalu bulan
    parsed.sort(key=lambda x: sort_key_period(x["info"]))

    log.append(("📋", f"{len(parsed)} file akan digabung (sudah diurutkan)"))

    frames = []
    for item in parsed:
        f = item["file"]
        info = item["info"]
        try:
            f.seek(0)
            fname = f.name.lower()
            if fname.endswith(".csv"):
                df = pd.read_csv(f)
            elif fname.endswith(".xls"):
                df = pd.read_excel(f, engine="xlrd")
            else:
                df = pd.read_excel(f, engine="openpyxl")

            if add_source_col:
                df.insert(0, source_col_name, f.name)
            if add_month_col and info["label"]:
                df.insert(1 if add_source_col else 0, month_col_name, info["label"])

            frames.append({"df": df, "info": info, "name": f.name})
            log.append(("✅", f"{f.name} → {len(df):,} baris | Periode: {info['label']}"))
        except Exception as e:
            log.append(("❌", f"Gagal baca {f.name}: {e}"))

    if not frames:
        return None, log

    # Build Excel
    wb = Workbook()
    wb.remove(wb.active)

    h_fill   = PatternFill('solid', start_color='1F4E79')
    h_font   = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
    h_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    alt_fill = PatternFill('solid', start_color='EBF3FB')
    d_font   = Font(name='Calibri', size=9)
    d_align  = Alignment(vertical='center')
    thin     = Side(style='thin', color='B0C4DE')
    bdr      = Border(left=thin, right=thin, top=thin, bottom=thin)

    def write_df_to_sheet(ws, df):
        df_str = df.fillna('-')
        for ci, col in enumerate(df.columns, 1):
            c = ws.cell(row=1, column=ci, value=col)
            c.font = h_font; c.fill = h_fill
            c.alignment = h_align; c.border = bdr

        for ri, row in enumerate(df_str.itertuples(index=False), 2):
            fill = alt_fill if ri % 2 == 0 else PatternFill('solid', start_color='FFFFFF')
            for ci, val in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = d_font; c.fill = fill
                c.alignment = d_align; c.border = bdr

        for ci, col in enumerate(df.columns, 1):
            max_len = max(
                len(str(col)),
                df_str.iloc[:, ci-1].astype(str).str.len().max() if len(df_str) else 0
            )
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 3, 45)

        ws.freeze_panes = 'A2'
        ws.row_dimensions[1].height = 32

    if sheet_per_month:
        # Satu sheet per file/bulan
        for item in frames:
            label = item["info"]["label"] or item["name"].rsplit(".", 1)[0]
            sheet_title = re.sub(r'[\\/*?:\[\]]', '', label)[:31]
            ws = wb.create_sheet(title=sheet_title)
            write_df_to_sheet(ws, item["df"])
        log.append(("📑", f"Dibuat {len(frames)} sheet (satu per bulan/file)"))

        # Sheet gabungan semua
        all_df = pd.concat([i["df"] for i in frames], ignore_index=True)
        ws_all = wb.create_sheet(title="SEMUA DATA")
        write_df_to_sheet(ws_all, all_df)
        log.append(("📊", f"Sheet 'SEMUA DATA' → {len(all_df):,} total baris"))

    else:
        # Satu sheet gabungan
        all_df = pd.concat([i["df"] for i in frames], ignore_index=True)
        ws = wb.create_sheet(title="Data Gabungan")
        write_df_to_sheet(ws, all_df)
        log.append(("📊", f"Total gabungan → {len(all_df):,} baris"))

    # Summary sheet
    if add_summary_sheet:
        ws_info = wb.create_sheet(title="Ringkasan")

        title_font = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
        title_fill = PatternFill('solid', start_color='1F4E79')
        sub_font   = Font(bold=True, name='Calibri', size=10)
        sub_fill   = PatternFill('solid', start_color='BDD7EE')
        data_font  = Font(name='Calibri', size=9)

        ws_info.merge_cells('A1:D1')
        c = ws_info['A1']
        c.value = "Ringkasan Penggabungan File — RME Data Cleaner"
        c.font = title_font; c.fill = title_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws_info.row_dimensions[1].height = 24

        headers = ["No", "Nama File", "Periode", "Jumlah Baris"]
        for ci, h in enumerate(headers, 1):
            c = ws_info.cell(row=2, column=ci, value=h)
            c.font = sub_font; c.fill = sub_fill
            c.alignment = Alignment(horizontal='center')

        total_rows = 0
        for idx, item in enumerate(frames, 1):
            ws_info.cell(row=idx+2, column=1, value=idx).font = data_font
            ws_info.cell(row=idx+2, column=2, value=item["name"]).font = data_font
            ws_info.cell(row=idx+2, column=3, value=item["info"]["label"]).font = data_font
            ws_info.cell(row=idx+2, column=4, value=len(item["df"])).font = data_font
            total_rows += len(item["df"])

        last_row = len(frames) + 3
        c_total = ws_info.cell(row=last_row, column=3, value="TOTAL")
        c_total.font = Font(bold=True, name='Calibri', size=9)
        c_val = ws_info.cell(row=last_row, column=4, value=total_rows)
        c_val.font = Font(bold=True, name='Calibri', size=9)

        ws_info.column_dimensions['A'].width = 5
        ws_info.column_dimensions['B'].width = 40
        ws_info.column_dimensions['C'].width = 20
        ws_info.column_dimensions['D'].width = 15

        log.append(("📃", "Sheet 'Ringkasan' berhasil dibuat"))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), log


# ─────────────────────────────────────────────
#  SIDEBAR
# ─────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div style="padding: 1rem 0 1.5rem;">
        <div style="font-size:1.5rem; font-weight:800; color:#f0f9ff; letter-spacing:-0.03em;">⚕️ RME Cleaner</div>
        <div style="font-size:0.75rem; color:#475569; margin-top:0.2rem;">Data Quality Tool v2.0</div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("**📂 Upload File**")
    uploaded_file = st.file_uploader(
        "Pilih file Excel atau CSV (1 file untuk clean)",
        type=["xlsx", "xls", "csv"],
        label_visibility="collapsed"
    )

    st.markdown("---")
    st.markdown("**⚙️ Opsi Pembersihan**")

    opt_remove_dup   = st.checkbox("Hapus baris duplikat", value=True)
    opt_strip_comma  = st.checkbox("Hapus trailing koma/spasi", value=True)
    opt_strip_spaces = st.checkbox("Hapus spasi dalam NIK & No Penjamin", value=True)
    opt_format_dates = st.checkbox("Format tanggal ke dd/mm/yyyy", value=True)
    opt_std_ket      = st.checkbox("Standarisasi kolom Keterangan", value=True)
    opt_upper_ket    = st.checkbox("↳ UPPERCASE Keterangan", value=True)
    opt_upper_nama   = st.checkbox("UPPERCASE kolom Nama", value=True)
    opt_title_desa   = st.checkbox("Title Case kolom Desa", value=True)

    st.markdown("---")
    st.markdown("**🔄 Nilai Pengganti**")
    fill_keterangan = st.text_input("Keterangan kosong/simbol", value="Tidak Ada Keterangan")
    fill_pekerjaan  = st.text_input("Pekerjaan kosong", value="Tidak Diketahui")
    fill_rm_lama    = st.text_input("RM Lama kosong", value="-")
    fill_desa       = st.text_input("Desa kosong", value="-")

    st.markdown("---")
    st.markdown("**📤 Format Output**")
    output_format = st.selectbox("Format download", ["Excel (.xlsx)", "CSV (.csv)"])
    sheet_name    = st.text_input("Nama sheet (Excel)", value="Data Bersih")


# ─────────────────────────────────────────────
#  MAIN HEADER
# ─────────────────────────────────────────────
st.markdown("""
<div class="main-header">
    <h1>🏥 RME Data Cleaner</h1>
    <p>Pembersihan & Penggabungan otomatis data Rekam Medis Elektronik Puskesmas</p>
</div>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────
#  TABS: Clean vs Merge
# ─────────────────────────────────────────────
tab_clean, tab_merge = st.tabs(["🧹 Bersihkan Data", "🔗 Gabung File Excel"])


# ════════════════════════════════════════════
#  TAB 1 — CLEAN
# ════════════════════════════════════════════
with tab_clean:

    if uploaded_file is None:
        st.markdown("""
        <div class="upload-hint">
            <div style="font-size:2.5rem; margin-bottom:0.75rem;">📂</div>
            <div style="color:#7dd3fc; font-weight:600; font-size:1rem; margin-bottom:0.4rem;">
                Upload file Excel atau CSV di sidebar kiri
            </div>
            <div style="font-size:0.8rem; color:#475569;">
                Format yang didukung: .xlsx · .xls · .csv<br>
                Data RME Puskesmas (Januari–Desember)
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown('<div class="section-card"><div class="section-title">📋 Fitur Pembersihan</div>', unsafe_allow_html=True)
        cols = st.columns(3)
        features = [
            ("🗑️", "Hapus Duplikat", "Deteksi dan hapus baris yang persis sama"),
            ("✂️", "Trailing Koma", "Hapus karakter ' ,' di akhir sel"),
            ("📅", "Format Tanggal", "Ubah ke format dd/mm/yyyy"),
            ("📝", "Standarisasi Keterangan", "Ubah '-' dan '=' menjadi teks bermakna"),
            ("🔠", "Normalisasi Teks", "UPPERCASE nama, Title Case desa"),
            ("📋", "Isi Nilai Kosong", "Isi kolom kosong dengan nilai default"),
        ]
        for i, (icon, title, desc) in enumerate(features):
            with cols[i % 3]:
                st.markdown(f"""
                <div style="background:rgba(30,41,59,0.6); border:1px solid rgba(99,179,237,0.15);
                            border-radius:10px; padding:1rem; margin-bottom:0.75rem; text-align:center;">
                    <div style="font-size:1.8rem;">{icon}</div>
                    <div style="color:#e2e8f0; font-weight:700; font-size:0.9rem; margin:0.4rem 0 0.3rem;">{title}</div>
                    <div style="color:#64748b; font-size:0.78rem;">{desc}</div>
                </div>
                """, unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

    else:
        @st.cache_data
        def load_data(file):
            name = file.name.lower()
            if name.endswith(".csv"):
                return pd.read_csv(file), ["Sheet1"]
            elif name.endswith(".xls"):
                xf = pd.ExcelFile(file, engine="xlrd")
                return pd.read_excel(file, sheet_name=xf.sheet_names[0], engine="xlrd"), xf.sheet_names
            else:
                xf = pd.ExcelFile(file, engine="openpyxl")
                return pd.read_excel(file, sheet_name=xf.sheet_names[0], engine="openpyxl"), xf.sheet_names

        try:
            df_raw, sheet_names = load_data(uploaded_file)
        except Exception as e:
            st.error(f"❌ Gagal membaca file: {e}")
            st.stop()

        issues = detect_issues(df_raw)
        total_nulls = sum(issues.get("nulls", {}).values())
        dup_count   = issues.get("duplicates", 0)

        st.markdown(f"""
        <div class="metric-row">
            <div class="metric-card">
                <div class="label">Total Baris</div>
                <div class="value">{len(df_raw):,}</div>
            </div>
            <div class="metric-card">
                <div class="label">Total Kolom</div>
                <div class="value">{len(df_raw.columns)}</div>
            </div>
            <div class="metric-card">
                <div class="label">Duplikat</div>
                <div class="value {'bad' if dup_count else 'good'}">{dup_count}</div>
            </div>
            <div class="metric-card">
                <div class="label">Nilai Kosong</div>
                <div class="value {'bad' if total_nulls else 'good'}">{total_nulls:,}</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        col_left, col_right = st.columns([1, 2])

        with col_left:
            st.markdown('<div class="section-card">', unsafe_allow_html=True)
            st.markdown('<div class="section-title">🔍 Masalah Terdeteksi</div>', unsafe_allow_html=True)

            if not issues:
                st.markdown('<div class="log-item"><span class="log-icon">✅</span> Data sudah bersih!</div>', unsafe_allow_html=True)
            else:
                if "duplicates" in issues:
                    st.markdown(f"""<div class="log-item"><span class="log-icon">🔴</span>
                    <span><b style="color:#f87171">{issues['duplicates']} baris duplikat</b></span></div>""", unsafe_allow_html=True)
                if "trailing_comma" in issues:
                    cols_str = ", ".join(issues["trailing_comma"][:4])
                    st.markdown(f"""<div class="log-item"><span class="log-icon">🟡</span>
                    <span>Trailing koma:<br><span class="badge badge-yellow">{cols_str}</span></span></div>""", unsafe_allow_html=True)
                if "nulls" in issues:
                    null_items = "".join([f'<span class="badge badge-red" style="margin:2px;">{c}: {v}</span> '
                        for c, v in list(issues["nulls"].items())[:8]])
                    st.markdown(f"""<div class="log-item"><span class="log-icon">🔴</span>
                    <span>Nilai kosong:<br>{null_items}</span></div>""", unsafe_allow_html=True)
                if "placeholders" in issues:
                    st.markdown(f"""<div class="log-item"><span class="log-icon">🟡</span>
                    <span>Placeholder: <span class="badge badge-yellow">{', '.join(issues['placeholders'][:4])}</span></span></div>""", unsafe_allow_html=True)
                if "date_cols" in issues:
                    st.markdown(f"""<div class="log-item"><span class="log-icon">🔵</span>
                    <span>Kolom tanggal: <span class="badge badge-blue">{', '.join(issues['date_cols'])}</span></span></div>""", unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)

            st.markdown('<div class="section-card">', unsafe_allow_html=True)
            st.markdown('<div class="section-title">🗂️ Pilih Kolom</div>', unsafe_allow_html=True)
            cols_to_drop = st.multiselect("Kolom yang ingin dihapus", options=list(df_raw.columns), default=[], placeholder="Pilih kolom...")
            sort_col = st.selectbox("Urutkan berdasarkan", options=["(tidak diurutkan)"] + list(df_raw.columns), index=0)
            st.markdown('</div>', unsafe_allow_html=True)

        with col_right:
            st.markdown('<div class="section-card">', unsafe_allow_html=True)
            st.markdown('<div class="section-title">👁️ Preview Data Asli (10 baris pertama)</div>', unsafe_allow_html=True)
            st.dataframe(df_raw.head(10), use_container_width=True, height=280)
            st.markdown('</div>', unsafe_allow_html=True)

        st.markdown("---")
        btn_col, _ = st.columns([1, 3])
        with btn_col:
            run_clean = st.button("🚀 Bersihkan Data Sekarang", use_container_width=True)

        if run_clean or "df_clean" in st.session_state:
            if run_clean:
                strip_space_cols = []
                for col in df_raw.select_dtypes(include="object").columns:
                    if "nik" in col.lower() or "penjamin" in col.lower():
                        strip_space_cols.append(col)

                fill_nulls_map = {}
                if fill_pekerjaan and "Pekerjaan" in df_raw.columns:
                    fill_nulls_map["Pekerjaan"] = fill_pekerjaan
                if fill_rm_lama and "RM Lama" in df_raw.columns:
                    fill_nulls_map["RM Lama"] = fill_rm_lama
                if fill_desa and "Desa" in df_raw.columns:
                    fill_nulls_map["Desa"] = fill_desa

                rename_map = {}
                if "kategori" in df_raw.columns:
                    rename_map["kategori"] = "Kategori"

                options = {
                    "remove_duplicates":     opt_remove_dup,
                    "strip_trailing_comma":  opt_strip_comma,
                    "strip_spaces":          opt_strip_spaces,
                    "strip_spaces_cols":     strip_space_cols,
                    "format_dates":          opt_format_dates,
                    "date_format":           "%d/%m/%Y",
                    "standardize_keterangan":opt_std_ket,
                    "placeholder_replacement": fill_keterangan,
                    "uppercase_keterangan":  opt_upper_ket,
                    "uppercase_nama":        opt_upper_nama,
                    "titlecase_desa":        opt_title_desa,
                    "fill_nulls":            fill_nulls_map,
                    "rename_cols":           rename_map,
                    "drop_cols":             cols_to_drop,
                    "sort_by":               sort_col if sort_col != "(tidak diurutkan)" else None,
                }

                with st.spinner("Membersihkan data..."):
                    df_clean, clean_log = clean_dataframe(df_raw, options)
                st.session_state["df_clean"] = df_clean
                st.session_state["clean_log"] = clean_log

            df_clean = st.session_state["df_clean"]
            clean_log = st.session_state.get("clean_log", [])

            removed_rows = len(df_raw) - len(df_clean)
            st.markdown(f"""
            <div class="metric-row">
                <div class="metric-card"><div class="label">Baris Tersisa</div>
                <div class="value good">{len(df_clean):,}</div></div>
                <div class="metric-card"><div class="label">Baris Dihapus</div>
                <div class="value {'bad' if removed_rows else 'good'}">{removed_rows}</div></div>
                <div class="metric-card"><div class="label">Kolom Tersisa</div>
                <div class="value">{len(df_clean.columns)}</div></div>
                <div class="metric-card"><div class="label">Nilai Kosong</div>
                <div class="value good">{int(df_clean.isnull().sum().sum())}</div></div>
            </div>
            """, unsafe_allow_html=True)

            res_left, res_right = st.columns([1, 2])
            with res_left:
                st.markdown('<div class="section-card">', unsafe_allow_html=True)
                st.markdown('<div class="section-title">✅ Log Pembersihan</div>', unsafe_allow_html=True)
                for icon, msg in clean_log:
                    st.markdown(f'<div class="log-item"><span class="log-icon">{icon}</span><span>{msg}</span></div>', unsafe_allow_html=True)
                st.markdown('</div>', unsafe_allow_html=True)

            with res_right:
                st.markdown('<div class="section-card">', unsafe_allow_html=True)
                st.markdown('<div class="section-title">✨ Preview Data Bersih</div>', unsafe_allow_html=True)
                st.dataframe(df_clean.head(15), use_container_width=True, height=320)
                st.markdown('</div>', unsafe_allow_html=True)

            st.markdown("---")
            dl_col1, dl_col2, dl_col3 = st.columns([1, 1, 2])
            base_name = uploaded_file.name.rsplit(".", 1)[0]

            if output_format == "Excel (.xlsx)":
                excel_bytes = to_excel_bytes(df_clean, sheet_name)
                with dl_col1:
                    st.download_button("⬇️ Download Excel (.xlsx)", data=excel_bytes,
                        file_name=f"{base_name}_cleaned.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True)
            else:
                csv_bytes = df_clean.fillna('-').to_csv(index=False).encode("utf-8-sig")
                with dl_col1:
                    st.download_button("⬇️ Download CSV", data=csv_bytes,
                        file_name=f"{base_name}_cleaned.csv", mime="text/csv",
                        use_container_width=True)

            with dl_col2:
                csv_bytes2 = df_clean.fillna('-').to_csv(index=False).encode("utf-8-sig")
                if output_format == "Excel (.xlsx)":
                    st.download_button("⬇️ Download CSV", data=csv_bytes2,
                        file_name=f"{base_name}_cleaned.csv", mime="text/csv",
                        use_container_width=True)

            remaining_nulls = df_clean.isnull().sum()
            remaining_nulls = remaining_nulls[remaining_nulls > 0]
            if not remaining_nulls.empty:
                with st.expander("⚠️ Nilai kosong yang masih tersisa"):
                    st.dataframe(remaining_nulls.reset_index().rename(columns={"index": "Kolom", 0: "Jumlah Kosong"}),
                        use_container_width=True)


# ════════════════════════════════════════════
#  TAB 2 — MERGE
# ════════════════════════════════════════════
with tab_merge:

    st.markdown("""
    <div class="merge-card">
        <div class="merge-title">🔗 Gabungkan File Excel/CSV Berdasarkan Periode</div>
        <div style="color:#94a3b8; font-size:0.85rem; line-height:1.6;">
            Upload beberapa file sekaligus (Januari–Desember atau antar tahun).
            File akan <b style="color:#c084fc;">diurutkan otomatis</b> berdasarkan bulan/tahun yang terdeteksi dari nama file.<br>
            <span style="color:#64748b; font-size:0.78rem;">
            Contoh nama: <code>rme_januari_2024.xlsx</code> · <code>data_feb_2025.xlsx</code> · <code>laporan_03_2024.xlsx</code>
            </span>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # Upload multiple files
    merge_files = st.file_uploader(
        "Upload file untuk digabungkan (bisa pilih banyak sekaligus)",
        type=["xlsx", "xls", "csv"],
        accept_multiple_files=True,
        key="merge_uploader",
        label_visibility="collapsed"
    )

    if not merge_files:
        st.markdown("""
        <div class="upload-hint">
            <div style="font-size:2.5rem; margin-bottom:0.75rem;">📁</div>
            <div style="color:#c084fc; font-weight:600; font-size:1rem; margin-bottom:0.4rem;">
                Upload 2 atau lebih file Excel/CSV di atas
            </div>
            <div style="font-size:0.8rem; color:#475569;">
                Sistem akan mendeteksi bulan & tahun dari nama file secara otomatis<br>
                lalu mengurutkannya sebelum digabungkan
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("""
        <div class="section-card">
            <div class="section-title">💡 Format Nama File yang Didukung</div>
        """, unsafe_allow_html=True)

        ex_col1, ex_col2 = st.columns(2)
        with ex_col1:
            st.markdown("""
            <div style="color:#94a3b8; font-size:0.83rem; line-height:2;">
                <b style="color:#c084fc;">Nama Bulan (Indonesia/Inggris)</b><br>
                📄 <code>rme_januari_2024.xlsx</code><br>
                📄 <code>data_february_2025.xlsx</code><br>
                📄 <code>laporan-maret-2024.xlsx</code><br>
                📄 <code>pasien_april2025.xlsx</code>
            </div>
            """, unsafe_allow_html=True)
        with ex_col2:
            st.markdown("""
            <div style="color:#94a3b8; font-size:0.83rem; line-height:2;">
                <b style="color:#c084fc;">Angka Bulan</b><br>
                📄 <code>rme_01_2024.xlsx</code><br>
                📄 <code>data_02_2025.xlsx</code><br>
                📄 <code>laporan_12_2024.xlsx</code><br>
                📄 <code>rekap_2024_03.xlsx</code>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("</div>", unsafe_allow_html=True)

    else:
        # Preview file detection
        st.markdown('<div class="section-card">', unsafe_allow_html=True)
        st.markdown('<div class="section-title">🔍 Deteksi Periode dari Nama File</div>', unsafe_allow_html=True)

        detected_periods = []
        for f in merge_files:
            info = extract_period_from_filename(f.name)
            detected_periods.append({"file": f, "info": info})

        detected_periods.sort(key=lambda x: sort_key_period(x["info"]))

        tags_html = ""
        for item in detected_periods:
            info = item["info"]
            badge_label = f"{info['label']}" if info["label"] else "?"
            tags_html += f'<span class="file-tag sorted">📄 {item["file"].name} → <b>{badge_label}</b></span>'

        st.markdown(f'<div style="margin-bottom:0.5rem;">{tags_html}</div>', unsafe_allow_html=True)
        st.markdown(f'<div style="color:#64748b; font-size:0.78rem; margin-top:0.5rem;">✅ {len(merge_files)} file terdeteksi — akan digabung sesuai urutan di atas</div>', unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

        # Merge options
        st.markdown('<div class="section-card">', unsafe_allow_html=True)
        st.markdown('<div class="section-title">⚙️ Opsi Penggabungan</div>', unsafe_allow_html=True)

        opt_col1, opt_col2 = st.columns(2)
        with opt_col1:
            merge_add_source  = st.checkbox("Tambah kolom nama file sumber", value=True)
            merge_add_month   = st.checkbox("Tambah kolom periode (bulan/tahun)", value=True)
            merge_sheet_month = st.checkbox("Buat sheet terpisah per bulan/file", value=False)
        with opt_col2:
            merge_add_summary = st.checkbox("Buat sheet ringkasan", value=True)
            if merge_add_source:
                source_col_name = st.text_input("Nama kolom sumber", value="Sumber File")
            else:
                source_col_name = "Sumber File"
            if merge_add_month:
                month_col_name = st.text_input("Nama kolom periode", value="Periode")
            else:
                month_col_name = "Periode"

        output_merge_name = st.text_input(
            "Nama file output",
            value="data_gabungan_RME.xlsx",
            help="Nama file Excel hasil penggabungan"
        )

        st.markdown('</div>', unsafe_allow_html=True)

        # Merge button
        merge_btn_col, _ = st.columns([1, 3])
        with merge_btn_col:
            run_merge = st.button("🔗 Gabungkan Sekarang", use_container_width=True)

        if run_merge or "merge_result" in st.session_state:
            if run_merge:
                with st.spinner("Menggabungkan file..."):
                    merged_bytes, merge_log = merge_excel_files(
                        uploaded_files   = merge_files,
                        add_source_col   = merge_add_source,
                        add_month_col    = merge_add_month,
                        source_col_name  = source_col_name,
                        month_col_name   = month_col_name,
                        sheet_per_month  = merge_sheet_month,
                        add_summary_sheet= merge_add_summary,
                    )
                st.session_state["merge_result"] = merged_bytes
                st.session_state["merge_log"] = merge_log
                st.session_state["merge_filename"] = output_merge_name

            merged_bytes = st.session_state["merge_result"]
            merge_log    = st.session_state.get("merge_log", [])
            out_filename = st.session_state.get("merge_filename", "data_gabungan_RME.xlsx")

            # Log
            st.markdown('<div class="section-card">', unsafe_allow_html=True)
            st.markdown('<div class="section-title">📋 Log Penggabungan</div>', unsafe_allow_html=True)
            for icon, msg in merge_log:
                st.markdown(f'<div class="log-item"><span class="log-icon">{icon}</span><span>{msg}</span></div>', unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)

            # Total metric
            total_merged_rows = sum(
                int(re.search(r'([\d,]+) baris', m).group(1).replace(',', ''))
                for i, m in merge_log if re.search(r'([\d,]+) baris', m)
                and '→' in m and 'Periode' in m
            ) if any('→' in m for _, m in merge_log) else 0

            if merged_bytes:
                st.markdown("---")
                dl_merge_col, _ = st.columns([1, 3])
                with dl_merge_col:
                    st.download_button(
                        label="⬇️ Download File Gabungan (.xlsx)",
                        data=merged_bytes,
                        file_name=out_filename if out_filename.endswith(".xlsx") else out_filename + ".xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
